from datetime import datetime
import openpyxl

container_file = openpyxl.load_workbook("konteineri.xlsx")
container_list = container_file["Sheet1"]

driver_appearance = {}
total_value_per_driver = {}

for target_row in range(2, container_list.max_row + 1):
    driver_name = container_list.cell(target_row, 2).value

    if driver_name in driver_appearance:
        current_qty = driver_appearance[driver_name]
        driver_appearance[driver_name] = current_qty + 1
    else:
        driver_appearance[driver_name] = 1
print(driver_appearance)

for target_row in range(2, container_list.max_row + 1):
    loading_d = container_list.cell(target_row, 1).value
    loading_date = str(datetime.date(loading_d))
#  loading_date_1 = (datetime.strptime(loading_date, "%d.%m.%Y"))
    container_qty = container_list.cell(target_row, 3).value
    cubic_amount = container_list.cell(target_row, 4).value

    if loading_date in total_value_per_driver:
        current_total_value = total_value_per_driver.get(loading_date)
        total_value_per_driver[loading_date] = current_total_value + container_qty * cubic_amount
    else:
        total_value_per_driver[loading_date] = container_qty * cubic_amount
print(total_value_per_driver)

'''
for target_row in range(2, container_list.max_row + 1):
    driver_name = container_list.cell(target_row, 2).value

    if driver_name in container_per_date:
        current_qty = container_per_date[driver_name]
        container_per_date[driver_name] = current_qty + 1
    else:
        container_per_date[driver_name] = 1
print(container_per_date)
'''